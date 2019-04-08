from os import listdir
from openpyxl import load_workbook, Workbook
from openpyxl.styles import PatternFill, Color
from myLib import check_Inkwd, check_outkwd

files = listdir(".")
src_xlsx = load_workbook("src.xlsx")
src_sheet = src_xlsx.active

src_sheet.delete_cols(1) # Post Reaction
src_sheet.delete_cols(6,3) # Likes Count , Tagged User Count on This Comment, Tagged User Count in Entire Post
col_date = src_sheet['E'] # DATE COL

result_xlsx = Workbook()
result_sheet = result_xlsx.active

count = 0
#ID     CommentLink	Name	Date	Text
for row in src_sheet.iter_rows():

    cell_data = []
    # 첫 줄에 구분자 삽입
    if count == 0:
        cell_data.append("No")

        # 첫 줄(카테고리) 삽입
        for cell in row:
            cell_data.append(cell.value)
        result_sheet.append(cell_data)
        count = count + 1
        continue


    # 첫째칸이 비어있거나, Text가 비어있다면 패스
    if (row[0].value is None or row[4].value is None):
        continue

    # 이벤트 조건

    # 이벤트 조건과 맞을때 셀에 추가

    if(  check_Inkwd(row[4].value) and check_outkwd(row[4].value)):
        # 두번째 줄(A2) 부터 데이터 복사
        cell_data.append(count)
        count = count + 1

        for cell in row:
            if (cell.value is None):
                continue
            cell_data.append(cell.value)

        result_sheet.append(cell_data)



result_xlsx.save("result.xlsx")