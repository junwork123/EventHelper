from os import listdir
from openpyxl import load_workbook, Workbook
from Event import *
from openpyxl.styles import PatternFill, Color, Border, Side, Alignment # 셀 스타일 지정

OUT_KEYWORDS = []
IN_KEYWORDS = []

# 이벤트 옵션을 전역변수에 세팅
def setOptions(myEvent):
    OUT_KEYWORDS.extend(myEvent.getOpt_Cherry())
    IN_KEYWORDS.extend(myEvent.getOpt_URL())
    IN_KEYWORDS.extend(myEvent.getOpt_Keywords())

# 이벤트 조건과 맞을때 셀에 추가
def check_Keyword(value):

    for word in IN_KEYWORDS:
        if (not (word in value)):
            break

        if (word is IN_KEYWORDS[-1]):
            return True

    return False

# 체리피커를 걸러내는 함수
def check_Cherry(value):

    # 체리피커 키워드가 있을때 거른다
    for word in OUT_KEYWORDS:
        if (word in value):
            return False

    return True

def set_border(ws, min_row, max_row, min_col, max_col):
    rows = list(ws.iter_rows(min_row, max_row, min_col, max_col))
    side = Side(border_style='thin', color="FF000000")

    rows = list(rows) # we convert iterator to list for simplicity, but it's not memory efficient solution
    max_y = len(rows) - 1 # index of the last row
    for pos_y, cells in enumerate(rows):
        max_x = len(cells) - 1 # index of the last cell
    for pos_x, cell in enumerate(cells):
        border = Border(left=cell.border.left,
                        right=cell.border.right,
                        top=cell.border.top,
                        bottom=cell.border.bottom)
    if pos_x == 0:
        border.left = side
    if pos_x == max_x:
        border.right = side
    if pos_y == 0:
        border.top = side
    if pos_y == max_y:
        border.bottom = side

    # set new border only if it's one of the edge cells
    if pos_x == 0 or pos_x == max_x or pos_y == 0 or pos_y == max_y:
        cell.border = border

def copyRange(startCol, startRow, endCol, endRow, sheet):
    rangeSelected = []
    # Loops through selected Rows
    for i in range(startRow, endRow + 1, 1):
        # Appends the row to a RowSelected list
        rowSelected = []
        for j in range(startCol, endCol + 1, 1):
            rowSelected.append(sheet.cell(row=i, column=j).value)
        # Adds the RowSelected List and nests inside the rangeSelected
        rangeSelected.append(rowSelected)

    return rangeSelected

def pasteRange(startCol, startRow, endCol, endRow, sheetReceiving, copiedData):
    countRow = 0
    for i in range(startRow, endRow + 1, 1):
        countCol = 0
        for j in range(startCol, endCol + 1, 1):
            sheetReceiving.cell(row=i, column=j).value = copiedData[countRow][countCol]
            countCol += 1
        countRow += 1


def generateExcel(myEvent):

    # 매개변수로 받은 Event 인스턴스를 복사
    setOptions(myEvent)
    myAlignment = Alignment(horizontal = 'center', vertical = 'center')
    myBorder =  Border(left=Side(style='thin'),
                       right=Side(style='thin'),
                       top=Side(style='thin'),
                       bottom=Side(style='thin'))

    files = listdir(".")
    src_xlsx = load_workbook("excel\src.xlsx")
    src_sheet = src_xlsx.active

    src_sheet.delete_cols(8, 3)  # Tagged User Count on This Comment, Tagged User Count in Entire Post
    src_sheet.delete_cols(1, 3)  # Post Reaction, ID, Comment Link
    result_xlsx = Workbook()

    # 조건에 맞는 사람들을 뽑아내는 리스트 시트
    list_sheet = result_xlsx. active
    list_sheet['A1'] = "No"
    list_sheet['B1'] = "이름"
    list_sheet['C1'] = "날짜"
    list_sheet['D1'] = "댓글 내용"
    list_sheet['E1'] = "좋아요 수"

    # 추첨 시트 생성
    raffle_sheet = result_xlsx.create_sheet("추첨페이지")

    raffle_sheet['A1'] = "No"
    raffle_sheet['B1'] = "이름"
    raffle_sheet['C1'] = "난수값"
    raffle_sheet['D1'] = "순위"
    raffle_sheet['E1'] = "당첨자"
    raffle_sheet['F1'] = "상품"



    # 이벤트 조건에 맞는 사람들 추출
    validNum= 1
    for row in src_sheet.iter_rows():

        cell_data = []

        # 첫째칸이 비어있거나, Text가 비어있다면 패스
        if (row[0].value is None or row[2].value is None):
            continue

        # 이벤트 조건과 맞을때 셀에 추가
        if (check_Keyword(row[2].value) and check_Cherry(row[2].value)):
            # 두번째 줄(A2) 부터 데이터 복사
            cell_data.append(validNum)
            validNum = validNum + 1

            for cell in row:
                if (cell.value is None):
                    continue
                cell_data.append(cell.value)

            list_sheet.append(cell_data)

    # '추첨페이지'에서 추출한 리스트 중 당첨자 추첨
    # 'No.', '이름' 복사-붙여넣기
    copiedData = copyRange(1,2,2,validNum,list_sheet)
    pasteRange(1,2,2,validNum,raffle_sheet,copiedData)

    prizeDict = myEvent.getPrize()
    prizeDictkeys = myEvent.getPrize().keys()
    prizeNumLimit = 1
    for keys in prizeDict:
        print(prizeDict[keys])
        raffle_sheet.cell(row=prizeNumLimit+1, column=6).value = keys+"\n("+str(prizeDict[keys])+"명"+")"
        raffle_sheet.merge_cells(start_row=prizeNumLimit+1, start_column=6, end_row=prizeNumLimit + prizeDict[keys], end_column=6)

        prizeNumLimit = prizeNumLimit + prizeDict[keys]

    print(prizeNumLimit)

    for i in range(2, validNum):
        # '난수값' 열에 난수값 함수 삽입
        raffle_sheet.cell(row = i, column = 3).value = "=RAND()"
        # VLOOKUP 함수를 이용한 당첨자 추출
        if( i < prizeNumLimit+1 ):
            raffle_sheet.cell(row = i, column = 4).value = i - 1
            raffle_sheet.cell(row = i, column = 5).value = "=VLOOKUP(D"+str(i)+",$A$2:$C$"+str(validNum)+",2,0)"


    set_border(raffle_sheet, min_row = 1, max_row = prizeNumLimit, min_col = 4, max_col = 6 )
#청해부대

    # 최종 파일 저장
    result_xlsx.save("result.xlsx")